# -*- coding: utf-8 -*-
"""
Created on Tue Aug 14 09:55:15 2018

@author: sinra
"""
import os
import nltk
from nltk.tag import StanfordNERTagger
from nltk.tag import StanfordPOSTagger
from nltk.parse.stanford import StanfordParser
from nltk import Tree
from nltk.parse.stanford import StanfordDependencyParser
from stanfordcorenlp import StanfordCoreNLP
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.corpus import stopwords
from nltk.parse.corenlp import CoreNLPDependencyParser
from nltk.parse import DependencyGraph
from graphviz import Source
import re
from openpyxl import load_workbook
import joblib

os.environ['JAVA_HOME'] = "C:/Program Files/Java/jre1.8.0_181/bin/java.exe" 
os.environ['CLASSPATH'] = "D:/MIRLABWork/stanford-nlp/stanford-nltk/jar"
os.environ['STANFORD_MODELS'] = "D:/MIRLABWork/stanford-nlp/stanford-nltk/model"

wb = load_workbook(r"D:\MATLAB MFiles\Dr.Chen NER Project\PARSER TO BRAT\word_database_180716.xlsx")
print(wb.sheetnames)
sheet = wb.get_sheet_by_name("工作表1")

readYFC = open(r'D:\MATLAB MFiles\Dr.Chen NER Project\PARSER TO BRAT\test01modify.txt','r')
YFC_text =  readYFC.read()
readYFC.close()
joblib.dump(YFC_text,"Text")
#read keyword 
key_word = []
key_word_labels = []
for ii1 in sheet['A']:  
    key_word.append(ii1.value)
    
for ii2 in sheet['B']:
    key_word_labels.append(ii2.value)

key_word_all = {}
for ii3 in range(len(key_word)):
    key_word_all[key_word[ii3]] = key_word_labels[ii3]


YFC_text_token = nltk.word_tokenize(YFC_text)
stopword = [',',':','(',')']
stopword2 = ['.', ' ']
YFC_word_final = []
for ii5 in YFC_text_token:
    if ii5 not in stopword:
        YFC_word_final.append(ii5)

YFC_text_final = ' '.join(YFC_word_final)

#------------------------------------------------------------------------------#

YFC_text_sent_token = nltk.sent_tokenize(YFC_text_final)

zz3 = []
for i in range(0,5,1):
    zz3.append(nltk.sent_tokenize(YFC_text_final)[i])
YFC_text__to_temp = ' '.join(zz3)

zz3_5 = nltk.word_tokenize(YFC_text__to_temp)
zz4 = [ii6 for ii6 in zz3_5 if not ii6 in stopword2]
YFC_text__to_parse = ' '.join(zz4)
zzfinal = nltk.sent_tokenize(YFC_text__to_parse.lower())

eng_parser = StanfordDependencyParser()
res = list(eng_parser.parse(zzfinal))

dep_tree_dot_repr = [parse for parse in res][0].to_dot()
dtree = [parse for parse in res][0]
source = Source(dep_tree_dot_repr, filename="dep_tree", format="png")
source.view()

#------------------------------------------------------------------------------#

tt ={}
ttt4 = []
ttt4=res[0].nodes.keys()
for i in range(len(ttt4)+1):
    tt[i]=res[0].get_by_address(i)
    
kkw={}
for i in range(len(ttt4)+1):
    for j in range(len(key_word)):
        if key_word[j] == res[0].nodes[i]['word']:
            kkw2 = {}
            kkw2['labels'] = key_word_labels[j]
            kkw2['word'] = key_word[j] 
            kkw2['head'] = res[0].nodes[i]['head']
            kkw2['deps'] = res[0].nodes[i]['deps']
            kkw2['nodes'] = i
            kkw[i] = kkw2

finding_head = {}
implant_head = {}
location_head = {}

#------------------------------------------------------------------------------#

def trace_head_node(mun):
    word_node = []
    total_head_trace =[mun]
    head_trace = res[0].nodes[mun]['head']
    while head_trace != 0:
        total_head_trace.append(head_trace)
        head_trace = res[0].nodes[head_trace]['head']
    word_node = [kkw[mun]['word'],str(kkw[mun]['nodes'])]
    return word_node, total_head_trace

#------------------------------------------------------------------------------#
    
for i in kkw.keys():
    if kkw[i]['labels'] == 'finding':
        [word_node,total_head_trace] = trace_head_node(i)
        finding_head['_'.join(word_node)] = total_head_trace
        
    elif kkw[i]['labels'] == 'implant':
        [word_node, total_head_trace] = trace_head_node(i)
        implant_head['_'.join(word_node)] = total_head_trace
    else:
        [word_node, total_head_trace] = trace_head_node(i)
        location_head['_'.join(word_node)] = total_head_trace
    
dd = set(location_head['left_16'])
hh = set(finding_head['metastases_28'])

yy = dd ^ hh


print(dd)
print(hh)
print(yy)

    
    
